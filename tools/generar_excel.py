#!/usr/bin/env python3
"""
Tool: generar_excel
Crea archivos .xlsx con formato profesional a partir de datos tabulares.
Headers con fondo oscuro (#1F4E79), texto blanco bold, filas alternadas,
autoajuste de ancho de columna y formato numerico.

Uso: python generar_excel.py "ruta/de/salida.xlsx"
Input: JSON con headers y filas via env var EXCEL_DATA

Ejemplo de EXCEL_DATA:
{
  "headers": ["Departamento", "Unidades", "Valor COP"],
  "rows": [
    ["ANTIOQUIA", 33632, 126300000],
    ["BOGOTA", 16288, 78450000]
  ],
  "sheet_name": "Ventas",
  "title": "Reporte de Ventas por Departamento"
}
"""
import sys
import os
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


COLOR_HEADER_BG = '1F4E79'
COLOR_HEADER_FG = 'FFFFFF'
COLOR_ROW_ALT   = 'D6E4F0'
COLOR_TITLE_BG  = '2F5496'
COLOR_TITLE_FG  = 'FFFFFF'
COLOR_BORDER    = 'B0B0B0'

FONT_HEADER = Font(name='Calibri', bold=True, color=COLOR_HEADER_FG, size=11)
FONT_TITLE  = Font(name='Calibri', bold=True, color=COLOR_TITLE_FG, size=14)
FONT_DATA   = Font(name='Calibri', size=10)
FILL_HEADER = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type='solid')
FILL_TITLE  = PatternFill(start_color=COLOR_TITLE_BG, end_color=COLOR_TITLE_BG, fill_type='solid')
FILL_ALT    = PatternFill(start_color=COLOR_ROW_ALT, end_color=COLOR_ROW_ALT, fill_type='solid')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT   = Alignment(horizontal='left', vertical='center')
ALIGN_RIGHT  = Alignment(horizontal='right', vertical='center')
THIN_BORDER  = Border(
    left=Side(style='thin', color=COLOR_BORDER),
    right=Side(style='thin', color=COLOR_BORDER),
    top=Side(style='thin', color=COLOR_BORDER),
    bottom=Side(style='thin', color=COLOR_BORDER),
)


def _auto_width(headers, rows):
    col_widths = []
    for col_idx in range(len(headers)):
        max_len = len(str(headers[col_idx]))
        for row in rows:
            val = row[col_idx] if col_idx < len(row) else ''
            length = len(str(val))
            if isinstance(val, (int, float)):
                length = max(length, len(f'{val:,.0f}'))
            max_len = max(max_len, length)
        col_widths.append(min(max_len + 4, 40))
    return col_widths


def _is_numeric(val):
    if isinstance(val, (int, float)):
        return True
    if isinstance(val, str):
        try:
            float(val.replace(',', ''))
            return True
        except (ValueError, TypeError):
            return False
    return False


def _to_num(val):
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        try:
            cleaned = val.replace(',', '')
            if '.' in cleaned:
                return float(cleaned)
            return int(cleaned)
        except (ValueError, TypeError):
            return val
    return val


def generar_excel(
    headers: list,
    rows: list,
    output_path: str,
    sheet_name: str = 'Datos',
    title: str = None,
    column_widths: list = None,
) -> dict:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    current_row = 1

    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = FONT_TITLE
        title_cell.fill = FILL_TITLE
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 36
        current_row = 2

    header_row = current_row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 28

    data_start = header_row + 1
    for row_idx, row_data in enumerate(rows):
        excel_row = data_start + row_idx
        is_alt = row_idx % 2 == 1
        for col_idx, val in enumerate(row_data, start=1):
            val = _to_num(val)
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = FONT_DATA
            cell.border = THIN_BORDER
            if _is_numeric(val):
                cell.alignment = ALIGN_RIGHT
                if isinstance(val, int):
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = ALIGN_LEFT
            if is_alt:
                cell.fill = FILL_ALT

    widths = column_widths or _auto_width(headers, rows)
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    return {
        'success': True,
        'archivo': str(out_path.resolve()),
        'hojas': [sheet_name],
        'filas': len(rows),
        'columnas': len(headers),
    }


def main():
    raw = os.environ.get('EXCEL_DATA')
    if not raw:
        print(json.dumps({
            'success': False,
            'error': 'Variable de entorno EXCEL_DATA no encontrada. Pasa un JSON con headers, rows y output_path.',
        }))
        sys.exit(1)

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        print(json.dumps({'success': False, 'error': f'JSON invalido en EXCEL_DATA: {e}'}))
        sys.exit(1)

    headers = data.get('headers')
    rows = data.get('rows')
    if not headers or not rows:
        print(json.dumps({'success': False, 'error': 'EXCEL_DATA debe contener "headers" (list) y "rows" (list).'}))
        sys.exit(1)

    if len(sys.argv) > 1:
        output_path = sys.argv[1]
    else:
        output_path = data.get('output_path')
        if not output_path:
            print(json.dumps({'success': False, 'error': 'Debes proporcionar la ruta de salida como argumento o en EXCEL_DATA.output_path.'}))
            sys.exit(1)

    result = generar_excel(
        headers=headers,
        rows=rows,
        output_path=output_path,
        sheet_name=data.get('sheet_name', 'Datos'),
        title=data.get('title'),
        column_widths=data.get('column_widths'),
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
